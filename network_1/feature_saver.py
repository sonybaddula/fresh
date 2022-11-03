from openpyxl import Workbook, load_workbook
import scapy.all as scapy

# Creating the workbook and worksheet
wb = Workbook()
ws = wb.active

# Creating Column Titles
ws.cell(row=1, column=1, value="Prototype No")
ws.cell(row=1, column=2, value="Network No")
ws.cell(row=1, column=3, value="Network Name")
ws.cell(row=1, column=4, value="Feature No")
ws.cell(row=1, column=5, value="Attribute No")
ws.cell(row=1, column=6, value="Packet Type")
ws.cell(row=1, column=7, value="Packet Subtype")
ws.cell(row=1, column=8, value="Feature Value")
ws.cell(row=1, column=9, value="Security Protocol")
